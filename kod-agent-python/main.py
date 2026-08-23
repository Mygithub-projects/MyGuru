"""CLI servis AI e-KokoT6.
e-KokoT6 AI service CLI.

    python main.py semak            # semak konfigurasi, DB, dan fail sumber
    python main.py senarai          # senaraikan pelajar dari pangkalan data
    python main.py cerapan --ic 123456789012
"""

from __future__ import annotations

import argparse
import sys

from openpyxl import load_workbook

from ekoko_ai import cerapan as m_cerapan
from ekoko_ai import claude, db
from ekoko_ai.config import muat_tetapan

FAIL_SUMBER = ("namelist.xlsx", "GURU DATA.xlsx", "DUMM PAJSK.xlsx")


def cmd_semak(args: argparse.Namespace) -> int:
    """Semakan kesihatan — tidak memanggil API Claude.
    Health check — makes no Claude API call."""
    tetapan = muat_tetapan()
    print(f"Model      : {tetapan.ai_model}  (effort={tetapan.ai_effort})")
    print(f"Kunci API  : {'ditetapkan' if tetapan.anthropic_api_key else 'tiada — SDK akan cuba profil ant'}")
    print(f"DATA_DIR   : {tetapan.data_dir}")

    print("\nFail sumber:")
    for nama in FAIL_SUMBER:
        laluan = tetapan.data_dir / nama
        if not laluan.exists():
            print(f"  TIADA  {nama}")
            continue
        try:
            wb = load_workbook(laluan, read_only=True)
            ws = wb.worksheets[0]
            print(f"  OK     {nama}  (helaian '{ws.title}', {ws.max_row} baris)")
            wb.close()
        except Exception as e:  # noqa: BLE001 - laporkan apa sahaja yang gagal dibaca
            print(f"  RALAT  {nama}: {e}")

    print("\nPangkalan data:")
    try:
        with db.sambung(tetapan.database_url) as conn:
            for jadual, n in db.kiraan(conn).items():
                print(f"  {jadual:12} {n}")
    except Exception as e:  # noqa: BLE001 - sambungan gagal ialah keputusan yang sah
        print(f"  GAGAL menyambung: {e}")
        return 1
    return 0


def cmd_senarai(args: argparse.Namespace) -> int:
    tetapan = muat_tetapan()
    with db.sambung(tetapan.database_url) as conn:
        baris = db.senarai_pelajar(conn, had=args.had)
    if not baris:
        print("Tiada pelajar dalam pangkalan data.")
        return 0
    print(f"{'No. IC':<14} {'Kelas':<18} {'Markah':>7} {'Gred':>5}  Nama")
    for p in baris:
        markah = "-" if p["markahPajskT6"] is None else f"{p['markahPajskT6']:.1f}"
        print(f"{p['noIc']:<14} {(p['kelasT6'] or '-'):<18} {markah:>7} {(p['gredPajskT6'] or '-'):>5}  {p['nama']}")
    return 0


def cmd_cerapan(args: argparse.Namespace) -> int:
    tetapan = muat_tetapan()
    try:
        print(m_cerapan.jana(tetapan, args.ic))
    except LookupError as e:
        print(f"Ralat: {e}", file=sys.stderr)
        return 1
    except claude.RalatClaude as e:
        print(f"Ralat Claude: {e}", file=sys.stderr)
        return 1
    return 0


def bina_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="ekoko-ai", description=__doc__.splitlines()[0])
    sub = p.add_subparsers(dest="perintah", required=True)

    sub.add_parser("semak", help="semak konfigurasi, DB, dan fail sumber").set_defaults(fn=cmd_semak)

    s = sub.add_parser("senarai", help="senaraikan pelajar")
    s.add_argument("--had", type=int, default=10, help="bilangan baris (lalai 10)")
    s.set_defaults(fn=cmd_senarai)

    c = sub.add_parser("cerapan", help="jana cerapan PAJSK untuk seorang pelajar")
    c.add_argument("--ic", required=True, help="No. IC 12 digit")
    c.set_defaults(fn=cmd_cerapan)

    return p


def main() -> int:
    # Konsol Windows lalai kepada cp1252 dan akan merosakkan teks bukan-ASCII
    # (em-dash di sini, dan jawapan Bahasa Melayu daripada Claude).
    for aliran in (sys.stdout, sys.stderr):
        if hasattr(aliran, "reconfigure"):
            aliran.reconfigure(encoding="utf-8", errors="replace")

    args = bina_parser().parse_args()
    try:
        return args.fn(args)
    except RuntimeError as e:  # konfigurasi hilang, dsb.
        print(f"Ralat: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
